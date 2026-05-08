#!/usr/bin/env python3
"""
单人销售工作报告生成脚本
- 输入: Sales Excel 数据
- 过程: 按每个销售人员过滤名下数据（主表 + 工作记录），交给 Gemini 分析
- 输出: {销售名字}销售线索分析{年月}.pdf
"""

import os
import sys
import glob
import time
import json
import asyncio
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

from google import genai
from google.genai import types
from google.genai.errors import APIError

from person_ai_prompts import build_person_prompt

# ==================== 配置区 ====================
from dotenv import load_dotenv
load_dotenv()
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-2.5-flash"
os.environ["HTTP_PROXY"] = "http://127.0.0.1:7890"
os.environ["HTTPS_PROXY"] = "http://127.0.0.1:7890"

def load_org_config():
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception:
        config = {}
    org_chart = config.get("org_chart", {})
    region_map = {}
    for region, data in org_chart.items():
        if region == "支持部门": continue
        people = [data["manager"]["name"]]
        for sub in data.get("subordinates", []):
            people.append(sub["name"])
        region_map[region] = people
    
    lines = ["## 组织架构、汇报关系与 2026 年度销售目标"]
    for region, data in org_chart.items():
        mgr = data["manager"]
        postfix = "" if region == "支持部门" else "片区"
        
        region_quota_str = f" [片区总任务: {data['quota']}万]" if "quota" in data else ""
        mgr_quota_str = f" [个人年度任务: {mgr['quota']}万]" if "quota" in mgr else ""
        lines.append(f"- **{region}{postfix}{mgr['title']}**：{mgr['name']}{region_quota_str}{mgr_quota_str}")
        for sub in data.get("subordinates", []):
            sub_quota_str = f" [个人年度任务: {sub['quota']}万]" if "quota" in sub else ""
            lines.append(f"  - 下属{sub['title']}：{sub['name']}{sub_quota_str}")
    org_chart_md = "\n".join(lines)
    return region_map, org_chart_md

REGION_MAP, ORG_CHART_MD = load_org_config()

# 反向映射：人名 → 片区
PERSON_TO_REGION = {}
for region, people in REGION_MAP.items():
    for person in people:
        PERSON_TO_REGION[person] = region

ALL_SALES_PEOPLE = list(PERSON_TO_REGION.keys())

# ==================== 数据读取与筛选 ====================
def read_excel_data(excel_path):
    wb = load_workbook(excel_path)
    main_sheet = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in main_sheet[1]]
    
    main_data = []
    safe_to_original = {}
    import re
    for row in main_sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        main_data.append(record)
        # 建立 sheet name 映射表
        orig_name = str(record.get('项目名称', '')).strip()
        safe_name = re.sub(r'[\\\\*?:/\\[\\]]', '_', orig_name)[:31]
        safe_to_original[safe_name] = orig_name
    
    work_records = {}
    for sheet_name in wb.sheetnames[2:]:
        ws = wb[sheet_name]
        if ws.max_row and ws.max_row > 1:
            ws_headers = [cell.value for cell in ws[1]]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(ws_headers, row)))
            # 还原为原始项目名称，保证 AI 匹配一致性
            orig_name = safe_to_original.get(sheet_name, sheet_name)
            work_records[orig_name] = records
            
    return headers, main_data, work_records

def filter_data_by_person(main_data, work_records, person_name):
    # 筛选主报表
    person_data = []
    person_project_names = set()
    for record in main_data:
        responsible = str(record.get("负责人", ""))
        if person_name in responsible:
            person_data.append(record)
            proj_name = str(record.get("项目名称", "")).strip()
            if proj_name:
                person_project_names.add(proj_name)
    
    # 筛选工作记录
    person_work_records = {}
    for sheet_name, records in work_records.items():
        clean_sheet = sheet_name.strip()
        matched = False
        for proj_name in person_project_names:
            if clean_sheet in proj_name or proj_name[:28] in clean_sheet:
                matched = True
                break
        if matched:
            person_work_records[sheet_name] = records
    
    return person_data, person_work_records

# ==================== 格式化 ====================
def format_main_table(headers, data):
    key_cols = ['客\n户名称', '项目名称', '负责人', '商机阶段', '预期阶段',
                '项目金额', '预计开票日期', '预计开票金额']
    # fallbacks
    available_cols = [c for c in key_cols if c in headers]
    if not available_cols:
        available_cols = headers[:8]
    
    clean_cols = [str(c).replace('\n', '') for c in available_cols]
    lines = []
    lines.append('| ' + ' | '.join(clean_cols) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(available_cols)) + ' |')
    
    for record in data:
        vals = []
        for col in available_cols:
            v = record.get(col, '--')
            v = str(v).replace('\n', ' ').strip() if v else '--'
            vals.append(v)
        lines.append('| ' + ' | '.join(vals) + ' |')
    
    return '\n'.join(lines)

def format_work_records(work_records):
    lines = []
    for project, records in work_records.items():
        lines.append(f"\n### {project}")
        for r in records:
            content = str(r.get('记录内容', '--')).replace('\n', ' ')
            follow_type = str(r.get('跟进类型', '--'))
            date = str(r.get('日计划', '--'))
            lines.append(f"- [{date}] ({follow_type}) {content}")
    return '\n'.join(lines)

# ==================== Gemini API 调用 ====================
def call_gemini(prompt, task_name):
    import time
    for attempt in range(3):
        try:
            client = genai.Client(api_key=GEMINI_API_KEY, http_options={'api_version': 'v1', 'timeout': 600000})
            print(f"\n🤖 正在请求 Gemini ({GEMINI_MODEL}) 分析: {task_name} ...", flush=True)
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=prompt
            )
            result = response.text
            print(f"   ✅ {task_name} 分析完成 ({len(result)} 字)", flush=True)
            return result
        except Exception as e:
            err_str = str(e)
            print(f"   ❌ API 调用失败: {err_str}", flush=True)
            if any(k.lower() in err_str.lower() for k in ['429', 'RESOURCE_EXHAUSTED', 'quota', '503', 'UNAVAILABLE', 'high demand', 'SSL', 'EOF', 'ConnectError', 'Timeout', 'Connection', 'Proxy', 'RemoteProtocolError', 'disconnected', 'timed out']):
                wait = 60 * (attempt + 1)
                print(f"   ⏳ 触发自动重试机制，等待 {wait} 秒后重试 ({attempt+1}/3)...", flush=True)
                time.sleep(wait)
            else:
                raise
    raise Exception(f"Gemini API 多次重试后仍然失败")

# ==================== PDF 报告生成 ====================
def generate_person_report_pdf(person_name, analysis, pdf_path):
    import markdown
    import re
    
    today = datetime.now().strftime("%Y年%m月%d日")
    
    def md_to_html(md_text):
        """使用 markdown 库转换为 HTML，自动修复无空行或带缩进的表格"""
        lines = md_text.split('\\n')
        fixed_lines = []
        in_table = False
        for line in lines:
            stripped = line.strip()
            is_table_row = stripped.startswith('|') and stripped.count('|') >= 2
            
            if is_table_row:
                if not in_table:
                    if fixed_lines and fixed_lines[-1].strip() != '':
                        fixed_lines.append('')
                    in_table = True
                fixed_lines.append(line.lstrip(' \\t'))
            else:
                in_table = False
                fixed_lines.append(line)
                
        return markdown.markdown('\\n'.join(fixed_lines), extensions=['tables'])
    
    report_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
    @page {{ margin: 20mm 15mm; size: A4 portrait; }}
    * {{ box-sizing: border-box; }}
    body {{
        font-family: inherit;
        color: #1a1a1a;
        line-height: 1.7;
        margin: 0;
        padding: 0;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }}
    .cover {{
        text-align: center;
        padding-top: 250px;
        page-break-after: always;
    }}
    .cover h1 {{
        font-size: 38px;
        color: #1e40af;
        margin-bottom: 20px;
    }}
    .cover .subtitle {{
        font-size: 18px;
        color: #64748b;
        margin: 8px 0;
    }}
    .cover .divider {{
        width: 150px;
        height: 3px;
        background: linear-gradient(90deg, #3b82f6, #8b5cf6);
        margin: 40px auto;
        border-radius: 2px;
    }}
    .cover .meta {{
        font-size: 16px;
        color: #475569;
        margin-top: 100px;
    }}
    .content-page {{
        padding: 0 10px;
    }}
    h2, h3 {{ border-bottom: 2px solid #e2e8f0; padding-bottom: 15px; color: #1e40af; break-after: avoid; }}
    h2 {{ font-size: 24px; padding-left: 15px; border-left: 5px solid #3b82f6; border-bottom: 1px solid #e2e8f0; margin-top: 30px; }}
    table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; font-size: 11px; }}
    th, td {{ border: 1px solid #e2e8f0; padding: 6px 8px; text-align: left; }}
    th {{ background: #f8fafc; font-weight: bold; color: #475569; }}
    ul, ol {{ margin-bottom: 20px; padding-left: 20px; }}
    li {{ margin-bottom: 8px; }}
</style>
</head>
<body>
    <div class="cover">
        <h1>{person_name} 销售工作报告</h1>
        <div class="subtitle">线索与行动综合诊断书</div>
        <div class="divider"></div>
        <div class="meta">
            <p><strong>生成日期:</strong> {today}</p>
        </div>
    </div>
    <div class="content-page">
        {md_to_html(analysis)}
    </div>
</body>
</html>"""
    
    import tempfile
    from playwright.sync_api import sync_playwright
    
    tmp = tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8')
    tmp.write(report_html)
    tmp.close()
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f'file://{tmp.name}', wait_until='networkidle')
        page.wait_for_timeout(1000)
        
        page.pdf(
            path=pdf_path,
            format='A4',
            print_background=True,
            margin={'top': '15mm', 'bottom': '15mm', 'left': '12mm', 'right': '12mm'}
        )
        browser.close()
    
    os.unlink(tmp.name)

# ==================== 主流程 ====================
def find_latest_excel():
    downloads_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "downloads")
    if not os.path.exists(downloads_dir):
        return None
    files = glob.glob(os.path.join(downloads_dir, "**", "HAO_2026年*.xlsx"), recursive=True)
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def analyze_person():
    import argparse
    parser = argparse.ArgumentParser(description="单人销售报告生成脚本")
    parser.add_argument("excel_path", nargs="?", default=None, help="输入 Excel 文件路径：如果有最新则会自动查找")
    args = parser.parse_args()
    
    excel_path = args.excel_path
    if not excel_path:
        excel_path = find_latest_excel()
        if not excel_path:
            sys.exit(1)
            
    print("="*60, flush=True)
    print("🧠 个人销售 Review 分析系统 启动", flush=True)
    print("="*60, flush=True)
    
    headers, main_data, work_records = read_excel_data(excel_path)
    
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception:
        config = {}
    org_chart = config.get("org_chart", {})
    mgr_subs = {}
    for region, data in org_chart.items():
        mgr_name = data.get("manager", {}).get("name")
        subs = [s["name"] for s in data.get("subordinates", [])]
        if mgr_name:
            mgr_subs[mgr_name] = subs
    
    file_name = os.path.basename(excel_path)
    try:
        ymd = file_name.split("_")[2] # "20260321"
        yyyymm = ymd[:6]
    except Exception:
        yyyymm = datetime.now().strftime("%Y%m")
        
    for person_name in ALL_SALES_PEOPLE:
        if person_name == "周鑫": continue # 运营人员跳过
        
        person_data, person_records = filter_data_by_person(main_data, work_records, person_name)
        if not person_data:
            print(f"   ⏭️ 跳过 {person_name} (无销售数据)", flush=True)
            continue
            
        print(f"\n{'='*60}", flush=True)
        print(f"📍 开始处理: {person_name} (提取到 {len(person_data)} 条记录)", flush=True)
        print(f"{'='*60}", flush=True)
        
        person_table_md = format_main_table(headers, person_data)
        person_records_md = format_work_records(person_records)
        
        sub_table_md = ""
        if person_name in mgr_subs and mgr_subs[person_name]:
            all_sub_data = []
            for sub_name in mgr_subs[person_name]:
                s_data, _ = filter_data_by_person(main_data, work_records, sub_name)
                all_sub_data.extend(s_data)
            if all_sub_data:
                sub_table = format_main_table(headers, all_sub_data)
                sub_table_md = f"\n## 下属商机明细（只供片区经理计算团队总盘子和 KPI 缺口使用，无需逐个点评日志）\n{sub_table}\n"
        
        prompt = build_person_prompt(person_name, person_table_md, person_records_md, ORG_CHART_MD, sub_table_md)
        
        analysis = call_gemini(prompt, f"{person_name} 销售线索分析")
        
        region_name = PERSON_TO_REGION.get(person_name, "其它")
        region_dir = os.path.join(os.path.dirname(excel_path), region_name)
        os.makedirs(region_dir, exist_ok=True)
        pdf_path = os.path.join(region_dir, f"{person_name}销售线索分析{yyyymm}.pdf")
        
        print("🖨️ 正在渲染分析报告 PDF ...", flush=True)
        with open(pdf_path.replace('.pdf', '.md'), 'w', encoding='utf-8') as f:
            f.write(analysis)
        generate_person_report_pdf(person_name, analysis, pdf_path)
        
        if os.path.exists(pdf_path):
            size_mb = os.path.getsize(pdf_path) / (1024 * 1024)
            print(f"\n🎉🎉 {person_name} 报告 PDF 成功！")
            print(f"   📁 路径: {pdf_path}")
            print(f"   📏 大小: {size_mb:.1f}MB")
        else:
            print(f"   ❌ PDF 文件生成失败: {pdf_path}")

if __name__ == "__main__":
    analyze_person()
