import json
import time
import os
import sys
import pandas as pd
from playwright.sync_api import sync_playwright, expect
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

import shutil
import subprocess
import datetime

CONFIG_FILE = "config.json"
BASE_DOWNLOAD_DIR = os.path.join(os.getcwd(), "downloads")
BASE_BACKUP_DIR = os.path.join(os.getcwd(), "backup")

date_dirname = datetime.datetime.now().strftime("%Y-%m-%d")
DOWNLOAD_DIR = os.path.join(BASE_DOWNLOAD_DIR, date_dirname)
BACKUP_DIR = os.path.join(BASE_BACKUP_DIR, date_dirname)

os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# 启动时将已有的 downloads 文件移动到 backup
for filename in os.listdir(DOWNLOAD_DIR):
    file_path = os.path.join(DOWNLOAD_DIR, filename)
    if os.path.isfile(file_path):
        backup_path = os.path.join(BACKUP_DIR, filename)
        # 如果备份里有同名文件，为了防止覆盖报错，可以选择先删除或者加时间戳。这里简单选择强制覆盖
        if os.path.exists(backup_path):
            os.remove(backup_path)
        shutil.move(file_path, backup_path)

def load_config():
    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def run():
    config = load_config()
    with sync_playwright() as p:
        # 1. 启动浏览器
        browser = p.chromium.launch(headless=False, slow_mo=50) # 有头模式，放慢速度便于观察
        
        # ==========================================
        # 凭证状态持久化：绕过频繁的图形验证码
        # ==========================================
        state_file = "state.json"
        context_options = {
            "accept_downloads": True,
            "record_video_dir": DOWNLOAD_DIR
        }
        
        # 如果存在保存状态，说明之前成功登录过，就直接加载历史 Cookie
        if os.path.exists(state_file):
            context_options["storage_state"] = state_file
            print("====================================")
            print("✅ 检测到本地保存的登录凭证！开始【秒进】...")
            print("====================================")
            
        context = browser.new_context(**context_options)
        page = context.new_page()

        # 2. 检查凭证是否有效
        print(f"尝试加载目标报表: {config['report_url']}")
        page.goto(config['report_url'])
        
        # 留 3 秒用于判断是否被拦截并重定向回登录页
        page.wait_for_timeout(3000)
        
        if "login" in page.url.lower():
            print("凭证不存在或由于网站策略已失效。必须完成【单次】登录与人机验证...")
            page.goto(config['login_url'])
            
            # 等待 DOM 和基础渲染
            page.wait_for_load_state("domcontentloaded")
            page.wait_for_timeout(2000) 
            
            print(f"尝试切换到'账号登录'...")
            try:
                login_tab = page.locator("text=账号登录").first
                login_tab.wait_for(state="visible", timeout=5000)
                login_tab.click()
                page.wait_for_timeout(1000)
            except Exception: pass

            print("准备填写账号...")
            user_input = page.locator("input[placeholder*='手机'], input[placeholder*='邮箱'], input[placeholder*='账号'], input[name='username']").first
            try:
                user_input.wait_for(state="visible", timeout=5000)
                user_input.click() 
                user_input.fill(config["username"])
            except Exception: pass

            print("准备填写密码...")
            pwd_input = page.locator("input[type='password'], input[placeholder*='密码']").first
            try:
                pwd_input.wait_for(state="visible", timeout=5000)
                pwd_input.click()
                pwd_input.fill(config["password"])
            except Exception: pass
            
            print("准备勾选协议...")
            try:
                agreement_locator = page.locator(".protocol-text, label.el-checkbox, span:has-text('我已阅读并同意')").first
                agreement_locator.wait_for(state="attached", timeout=3000)
                agreement_locator.evaluate("node => node.click()")
            except Exception: pass
                
            print("准备点击登录按钮...")
            login_btn = page.locator("button:has-text('登录'), .login-btn, button[type='submit']").last
            try:
                login_btn.wait_for(state="visible", timeout=3000)
                login_btn.click()
            except:
                try: page.locator("text='登录'").last.click()
                except: pass

            print("\n=======================================================")
            print("🚨 首次启动必须处理图形验证码 🚨")
            print("请注意！！请在弹出的浏览器中【手动通过拼图/图形验证码】！")
            print("（后台每秒都会探测...只要一通过，就会自动把您的成果固化到本地！）")
            print("=======================================================\n")
            
            try:
                # 探测重定向以保存 cookie （这是防止下次弹验证码的核心）
                page.wait_for_url(lambda url: "login" not in url.lower() and "fxiaoke" in url.lower(), timeout=120000)
                context.storage_state(path=state_file)
                print("✅ 验证码通过！登录凭证已【成功固化】到本地 state.json！")
                print("以后只要凭证不过期，脚本就再也不会问您要验证码了~~")
            except Exception: pass
            
            print(f"\n跳回目标报表: {config['report_url']}")
            page.goto(config['report_url'])
            
        else:
            print("✅ 棒极了！历史免密凭证生效，【秒进】系统，无需任何输入直接接管！")
        
        page.wait_for_timeout(5000)

        # ==========================================
        # 无敌模式：暴力探测与真实物理点击器 (兼容动态 Canvas 图表层)
        # ==========================================
        def robust_click(selectors, name, max_waits=30):
            print(f"正在扫描并尝试真实物理点击: [{name}] ...")
            if isinstance(selectors, str):
                selectors = [selectors]
                
            for _ in range(max_waits):
                # 策略 0: 如果是在找"项目名称"(特征里含有 xkcharts)，主动先激起全局 Hover，逼图表层现原形！
                if "xkcharts" in str(selectors):
                    try:
                        # 闭着眼拿鼠标在报表区域疯狂划过，触发它按坐标按需生成的 tooltip
                        page.mouse.move(300, 300)
                        page.mouse.move(500, 500)
                        page.mouse.move(800, 600)
                    except: pass
                
                for sel in selectors:
                    visible_sel = f"{sel} >> visible=true"
                    
                    # 1. 搜主页面
                    try:
                        locs = page.locator(visible_sel).all()
                        for loc in locs:
                            try:
                                loc.hover(timeout=1000) 
                                loc.click(timeout=1000) 
                                print(f"   ✅ [真实鼠标反馈] 已成功点击 '{name}' (主页面)")
                                return True
                            except:
                                try:
                                    loc.click(force=True, timeout=1000)
                                    print(f"   ⚠️ [强压渗透] 已成功点击 '{name}' (主页面)")
                                    return True
                                except: pass
                    except: pass
                    
                    # 2. 搜所有 iframe
                    for frame in page.frames:
                        try:
                            locs = frame.locator(visible_sel).all()
                            for loc in locs:
                                try:
                                    loc.hover(timeout=1000)
                                    loc.click(timeout=1000)
                                    print(f"   ✅ [真实鼠标反馈] 已成功点击 '{name}' (内嵌框架)")
                                    return True
                                except:
                                    try:
                                        loc.click(force=True, timeout=1000)
                                        print(f"   ⚠️ [强压渗透] 已成功点击 '{name}' (内嵌框架)")
                                        return True
                                    except: pass
                        except: pass
                        
                page.wait_for_timeout(1000)
            print(f"   ❌ 严重警告：穷尽所有尝试后，未能在可见区域内触发 '{name}'")
            return False

        print("\n--- 开始执行报表业务操作 ---")
        
        if robust_click(["span.text:has-text('Hao')", "text='Hao'"], "报表文件夹 'Hao'", max_waits=20):
            page.wait_for_timeout(2000) # 让可能存在的展开动画跑完
            
        robust_click(["span.text:has-text('2026年大额开票机会统计')", "text='2026年大额开票机会统计'"], "具体报表 '2026年大额开票机会统计'", max_waits=20)
        
        print("报表开启，请等待数据网格渲染 (耐心等 8 秒)...")
        page.wait_for_timeout(8000)

        # ==========================================
        # 处理全自动报表配置与下载流程 (照着截图的参数)
        # ==========================================
        print("\n--- 准备接管导出流程 ---")
        try:
            # 放宽到 3 分钟等待 Fxiaoke 极其漫长的异步排队机制
            with page.expect_download(timeout=180000) as download_info:
                
                # 开始点击右上角大大的到处按钮
                export_selectors = [
                    "button[data-cy='bi-chartDesign-export']",
                    "button:has(span:has-text('导出'))", 
                    "button:has-text('导出')",
                    "text='导出'"
                ]
                if not robust_click(export_selectors, "主导出按钮", max_waits=20):
                    print("❌ 若未点到主导出按钮，后方的确认框将无法弹出！")
                    
                page.wait_for_timeout(2000)
                print("\n开始自动按照您截图的偏好配置弹窗属性...")
                
                # 这种带括号的往往因为 DOM 拆分或空格无法命中，所以取关键词 
                robust_click(["text='格式化报表'", "label:has-text('格式化报表')"], "1. 格式化报表选框", max_waits=3)
                robust_click(["text='CRM提醒'", "label:has-text('CRM提醒')"], "2. CRM提醒单选框", max_waits=3)
                robust_click(["text='包含筛选条件'", "label:has-text('包含筛选条件')"], "3. 包含筛选条件复选框", max_waits=3)
                robust_click(["button:has-text('确定')", "text='确定'"], "4. 最终[确定]按钮", max_waits=3)
                
                print("\n🚀 已推送请求给服务器！系统正在云端生成打包报表...")
                print("程序正在后台静默探测，等待纷享销客反馈'下载文件'弹窗...")
                
                download_trigger_clicked = False
                for _ in range(180): # 给它多一点处理时间，长达3分钟
                    for root in [page] + page.frames:
                        try:
                            # 尝试穿透所有的 frames 找这个下载对话框
                            dl_btns = root.locator("button:has-text('下载文件'), a:has-text('下载文件'), span:has-text('下载文件')").all()
                            for dl_btn in reversed(dl_btns):
                                if dl_btn.is_visible():
                                    dl_btn.evaluate("node => node.click()")
                                    download_trigger_clicked = True
                                    print("✅ 探测到了！并已核准点击最终的 '下载文件' 按钮！")
                                    break
                        except: pass
                        if download_trigger_clicked:
                            break
                    if download_trigger_clicked:
                        break
                    
                    page.wait_for_timeout(1000)
                    
                if not download_trigger_clicked:
                    print("⚠️ 仍未点到'下载文件'弹窗，您的账号有可能被限流排队，或者数据太大。")

            # 将捕获到的临时下载文件固化到我们要的目录里
            download = download_info.value
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            file_name = f"HAO_2026年大额开票机会统计_{timestamp}.xlsx"
            save_path = os.path.join(DOWNLOAD_DIR, file_name)
            
            print(f"\n⬇️ 正在将流数据提取落盘...")
            download.save_as(save_path)
            print(f"🎉🎉✅ 大功告成！全量报表已被自动接管并且保存至:\n {save_path}")
            
            # ==========================================
            # 第二阶段：读取 Excel，遍历项目获取截图并插入 Sheet
            # ==========================================
            print("\n" + "="*50)
            print("🚀 开始第二阶段：读取 Excel 提取项目名称，抓取工作记录截图并写入！")
            print("="*50)
            
            try:
                import pandas as pd
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image
                import re
                
                df = pd.read_excel(save_path)
                # 假设有一列叫"项目名称"，如果没有，找"项目"、"机会"等
                project_col = None
                for col in df.columns:
                    if "项目名称" in str(col) or "所属项目" in str(col) or "项目" in str(col):
                        project_col = col
                        break
                        
                if project_col is None:
                    print(f"❌ 在 Excel 中未找到类似 '项目名称' 的列，跳过截图分析步骤。可选列名有：{list(df.columns)}")
                else:
                    print(f"✅ 找到项目标识列：'{project_col}'，准备开始读取数据...")
                    
                    # 过滤掉空的、然后去重
                    projects = df[project_col].dropna().unique().tolist()
                    print(f"✅ 提取去重后的项目名称共 {len(projects)} 个，开启自动巡检...")
                    
                    # 重新刷新一次页面回归最外层清爽状体，或者只要重新点击报表名称也可以
                    page.goto(config['report_url'])
                    page.wait_for_timeout(5000)
                    if robust_click(["span.text:has-text('Hao')", "text='Hao'"], "重新点开文件夹 'Hao'", max_waits=5):
                        page.wait_for_timeout(2000)
                    robust_click(["span.text:has-text('2026年大额开票机会统计')", "text='2026年大额开票机会统计'"], "重新进入主报表列表", max_waits=5)
                    # 必须给足时间让首屏数据和Canvas层完全加载完毕，否则第一个项目必然会扫描失败被跳过
                    page.wait_for_timeout(8000)
                    
                    def xkcharts_canvas_radar_sweep(target_name):
                        print(f"📡 已启动高精度画布雷达，对图表渲染层进行悬停探测: {target_name} ...")
                        clean_target = target_name.replace('...', '').strip()
                        
                        canvases = page.locator("canvas, div[class*='xkcharts'], .xkcharts-table-body, .xk-table").all()
                        for chart in reversed(canvases):
                            try:
                                bbox = chart.bounding_box()
                                if not bbox or bbox['height'] < 50:
                                    continue
                                    
                                center_x = bbox['x'] + bbox['width'] / 2
                                center_y = bbox['y'] + bbox['height'] / 2
                                
                                page.mouse.move(0, 0)
                                page.wait_for_timeout(500)
                                
                                # 之前将提示框对象提出来全局复用 `.first` 是有风险的！
                                # 如果 Fxiaoke 渲染了多个不可见的缓存气泡，`.first` 永远去读那个死节点，导致一直错失真正的气泡！
                                # 现在改为动态遍历所有出现的气泡。
                                
                                # 尝试向下滚动最多 30 屏寻找 (大约需要 1-2 分钟)
                                prev_screen_tips = set()  # 用于检测连续两屏气泡相同（表示已到底）
                                for scroll_attempt in range(30):
                                    
                                    # =========== 开始对当前屏幕进行雷达光轴矩阵扫描 ===========
                                    # 【性能优化】从诊断日志确认：项目名称全部出现在 x=120，x=20 是客户名。
                                    # 只扫这两列，比之前扫 4 列快 2 倍。
                                    scan_x_list = [int(bbox['x']) + 20, int(bbox['x']) + 120]
                                    
                                    start_y = int(bbox['y']) + 50
                                    end_y = int(bbox['y'] + bbox['height']) - 5
                                    step_y = 12
                                    
                                    print(f"   [雷达扫描] 正在对第 {scroll_attempt+1} 屏进行高速扫描...")
                                    tips_seen_this_screen = set()
                                    for y in range(start_y, end_y, step_y):
                                        for x in scan_x_list:
                                            page.mouse.move(x, y)
                                            page.wait_for_timeout(150)  # 优化：300ms→150ms，诊断证明气泡响应速度很快
                                            
                                            try:
                                                all_tips = page.locator("div.xkcharts-ui-title-tip").all()
                                                for tip_el in all_tips:
                                                    try:
                                                        if tip_el.is_visible(timeout=30):
                                                            tip_text = tip_el.inner_text(timeout=30)
                                                            clean_tip = tip_text.replace('...', '').strip()
                                                            
                                                            if len(clean_tip) >= 3 and (clean_tip in clean_target or clean_target in clean_tip):
                                                                print(f"🎯 雷达锁定目标！坐标 ({x}, {y}) 成功命中气泡 '{tip_text}'！")
                                                                page.mouse.click(x, y)
                                                                return True
                                                            elif len(clean_tip) >= 2:
                                                                tips_seen_this_screen.add(clean_tip)
                                                    except:
                                                        pass
                                            except:
                                                pass
                                                    
                                    # ================= 当前屏幕未找到，触发向下滚动操作 =================
                                    # 【智能检测】如果连续两屏看到的气泡完全一样，说明列表已经到底，滚动无效，直接放弃
                                    if scroll_attempt > 0 and tips_seen_this_screen == prev_screen_tips:
                                        print(f"   ⚠️ [智能检测] 连续两屏气泡完全相同，列表已到底，停止无效滚动。")
                                        break
                                    prev_screen_tips = tips_seen_this_screen.copy()
                                    
                                    print(f"   [稳健滚动] 正在进行第 {scroll_attempt+1} 次下推...")
                                    try:
                                        page.mouse.wheel(0, 150)
                                        page.wait_for_timeout(200)
                                    except: pass
                                    
                                    # 等 Canvas 重绘，优化：1000ms→600ms
                                    page.wait_for_timeout(600)
                                    
                                    # 滚动后把鼠标拔出 Canvas 强制触发 onMouseOut 重置气泡引擎
                                    page.mouse.move(0, 0)
                                    page.wait_for_timeout(150)
                                    
                            except Exception as e:
                                print(f"   ⚠️ 画布外包裹解析跳过: {e}")
                                pass
                        # 所有屏（30屏）都滚完了还是没找到
                        return False
                    
                    wb = load_workbook(save_path)
                    
                    for proj_name in projects:
                        proj_name_str = str(proj_name).strip()
                        if not proj_name_str:
                            continue
                            
                        # 安全的工作表名称（Excel Sheet名字有特殊字符限制，且最长31个字符）
                        safe_sheet_name = re.sub(r'[\\*?:/\[\]]', '_', proj_name_str)[:31]
                        
                        if safe_sheet_name in wb.sheetnames:
                            print(f"⏩ 专属 Sheet 已存在，自动跳过去重: '{proj_name_str}'")
                            continue
                            
                        print(f"\n👉 正在处理详情: {proj_name_str} ...")
                        
                        # 在报表列表中寻找并点击对应的项目链接
                        # ===============================================
                        # 针对 xkcharts Canvas 报表独家抓包逻辑：
                        # 它画在 Canvas 上而不是 HTML，只能全屏扫描或者通过其底层虚拟化表格（Virtual Table）拿点击接口。
                        # 我们直接启动基于画布的高精度雷达来捕捉悬浮窗。
                        # ===============================================
                        
                        clicked_success = xkcharts_canvas_radar_sweep(proj_name_str)
                            
                        # 最强兜底判定：不管中间发什么了什么，哪怕是人工帮忙点开的
                        # 只要现在抽屉已经赫然在目了，我们就当它成功了！
                        if not clicked_success:
                            try:
                                drawer_check = page.locator(".el-drawer, .detail-container, .slide-panel, .crm-detail-drawer, [role='dialog']").last
                                if drawer_check.is_visible(timeout=500):
                                    print(f"✅ 检测到项目详情抽屉已自动/手动弹出！突破拦截框：{proj_name_str}")
                                    clicked_success = True
                            except: pass
                                
                        if clicked_success:
                            print(f"✅ 确认目标已打开，准备执行内部拉取：{proj_name_str}")
                            
                            # 等待侧滑或弹窗完全展开
                            page.wait_for_timeout(1000)
                            
                            #寻找“工作记录”页签并点击。由于抽屉内可能使用 Tabs 标签页，经常是 span 结构
                            work_record_selectors = [
                                "span:has-text('工作记录')",
                                "span.text:has-text('工作记录')",
                                "div.el-tabs__item:has-text('工作记录')",
                                ".tab-item:has-text('工作记录')",
                                "text='工作记录'"
                            ]
                            if robust_click(work_record_selectors, "工作记录页签", max_waits=5):
                                page.wait_for_timeout(2500) # 等待工作记录列表渲染，给予充足的网络拉取时间
                                
                                try:
                                    # 尝试解析并结构化提取工作记录表格
                                    print("正在深度解析底部工作记录表格流数据...")
                                    table_data = None
                                    js_extract_script = '''() => {
                                        // 只限制在可见区域查探，如果在 iframe 里，body 往往就是可视区
                                        const container = document.querySelector('.el-drawer, .detail-container, .slide-panel, .crm-detail-drawer, [role="dialog"]') || document.body;
                                        if (!container) return null;
                                        
                                        // 策略1：常见的高阶表格或原生表格
                                        const elTable = container.querySelector('.el-table, .xkcharts-table, table');
                                        if (elTable && elTable.tagName !== 'TABLE') {
                                            const ths = elTable.querySelectorAll('.el-table__header-wrapper th, thead th, .vxe-header--column');
                                            const headers = Array.from(ths).map(th => th.innerText.trim());
                                            const rows = Array.from(elTable.querySelectorAll('.el-table__body-wrapper tbody tr, tbody tr, .vxe-body--row')).map(tr => {
                                                return Array.from(tr.querySelectorAll('td')).map(td => td.innerText.trim());
                                            });
                                            if (headers.length > 0 || rows.length > 0) return {headers, rows};
                                        } else if (elTable && elTable.tagName === 'TABLE') {
                                            const headers = Array.from(elTable.querySelectorAll('th')).map(th => th.innerText.trim());
                                            const rows = Array.from(elTable.querySelectorAll('tbody tr')).map(tr => {
                                                return Array.from(tr.querySelectorAll('td, th')).map(td => td.innerText.trim());
                                            });
                                            if (headers.length > 0 || rows.length > 0) return {headers, rows};
                                        }
                                        
                                        // 策略2：纷享销客独有的 bi-list-table / 纯 div 格局
                                        const biTables = container.querySelectorAll('.bi-list-table, .crm-table, .list-table, div[class*="table-container"], .el-table__body-wrapper');
                                        for (const tbl of biTables) {
                                            // 找可能的表头
                                            const headerNodes = tbl.querySelectorAll('.bi-header-text, .th, .head-cell, [class*="header"] span');
                                            let headers = Array.from(headerNodes).map(h => h.innerText.trim()).filter(t => t);
                                            
                                            // 找主体行
                                            const bodyArea = tbl.querySelector('.bi-new-list-body, .crm-table-body, .tbody, [class*="body"]') || tbl;
                                            let rows = [];
                                            const rowNodes = bodyArea.querySelectorAll('.bi-list-item-container, .row, .tr, [class*="list-item"], [class*="table-row"], tr');
                                            // 避免选中子表格里的行，只找一级
                                            const targetRows = rowNodes.length > 0 ? Array.from(rowNodes).slice(0, 50) : Array.from(bodyArea.children).slice(0, 50);
                                            
                                            for (let r of targetRows) {
                                                const cellNodes = r.querySelectorAll('.bi-display-item, .td, .cell, [class*="cell"], td');
                                                if (cellNodes.length > 0) {
                                                    const rowData = Array.from(cellNodes).map(c => c.innerText.trim());
                                                    rows.push(rowData);
                                                } else {
                                                    const texts = r.innerText.trim().split('\\n').map(t => t.trim()).filter(t => t);
                                                    if (texts.length > 0) rows.push(texts);
                                                }
                                            }
                                            if (headers.length > 0 || rows.length > 0) return {headers, rows};
                                        }
                                        
                                        // 策略3：工作记录的 Feed 信息流模式（针对 .xxvui-fl-body.item-detail 等）
                                        let feedItems = container.querySelectorAll('.xxvui-fl-body.item-detail, .crm-feed-item, .record-item');
                                        if (feedItems.length === 0) {
                                            feedItems = document.querySelectorAll('.xxvui-fl-body.item-detail, .crm-feed-item, .record-item');
                                        }
                                        if (feedItems.length > 0) {
                                            const headers = ['记录内容', '跟进类型', '客户', '项目', '日计划'];
                                            let rows = [];
                                            for(let item of feedItems) {
                                                const texts = item.innerText.split('\\n').map(t => t.trim()).filter(t => t.length > 0 && t !== '复制' && t !== '查看更多' && !t.includes('CRM('));
                                                
                                                let rowDict = {'记录内容': texts[0] || ''};
                                                for(let i=0; i<texts.length-1; i++){
                                                    const key = texts[i];
                                                    if(['跟进类型', '客户', '项目', '日计划', '客户简称'].includes(key)) {
                                                        const val = texts[i+1];
                                                        rowDict[key] = val;
                                                    }
                                                }
                                                
                                                // 修复日计划可能粘连阅读数的bug
                                                let dateStr = rowDict['日计划'] || '';
                                                const match = dateStr.match(/\\d{4}-\\d{2}-\\d{2}/);
                                                if (match) {
                                                    dateStr = match[0];
                                                }
                                                
                                                rows.push([
                                                    rowDict['记录内容'],
                                                    rowDict['跟进类型'] || '',
                                                    rowDict['客户'] || rowDict['客户简称'] || '',
                                                    rowDict['项目'] || '',
                                                    dateStr
                                                ]);
                                            }
                                            if (rows.length > 0) return {headers, rows};
                                        }
                                        
                                        // 策略4：如果没有结构化特征，就用文本强行找（针对包含记录ID和跟进类型的地方）
                                        const tags = container.querySelectorAll('div, ul');
                                        for(let i = tags.length - 1; i>=0; i--){
                                            const txt = tags[i].innerText || '';
                                            if(txt.includes('记录ID') && txt.includes('跟进类型')) {
                                                const lines = txt.split('\\n').filter(l => l.trim().length > 0);
                                                // 粗略包装为数据行
                                                return {headers: ['提取信息'], rows: lines.map(l => [l])};
                                            }
                                        }
                                        
                                        return null;
                                    }'''
                                    
                                    for frame in [page] + page.frames:
                                        try:
                                            res = frame.evaluate(js_extract_script)
                                            if res and (res.get('headers') or res.get('rows')):
                                                table_data = res
                                                frame_id = getattr(frame, 'name', 'MainPage')
                                                if not frame_id: frame_id = getattr(frame, 'url', '')[:40]
                                                print(f"✅ 从 DOM 层级: '{frame_id}' 中成功抽出表格结构！")
                                                break
                                        except Exception as ex:
                                            # print(f"评价提取脚本时抛出异常: {ex}")
                                            pass
                                    
                                    if table_data and (table_data.get('headers') or table_data.get('rows')):
                                        print(f"✅ 成功抓取工作表格: 表头 {len(table_data['headers'])} 列, 数据 {len(table_data['rows'])} 行。")
                                        # 立即载入 Excel
                                        ws = wb.create_sheet(title=safe_sheet_name)
                                        # 写入表头，即使没抓到也可以给原图字段
                                        if table_data['headers']:
                                            ws.append(table_data['headers'])
                                        else:
                                            ws.append(['记录ID', '跟进类型', '客户简称', '项目', '日计划', '操作'])
                                            
                                        # 写入行
                                        for row in table_data['rows']:
                                            clean_row = [str(cell).replace('\n', ' ') for cell in row]
                                            ws.append(clean_row)
                                            
                                        # 调整列宽以及设置自动换行让单元格完整显示
                                        from openpyxl.styles import Alignment
                                        
                                        for col in ws.columns:
                                            max_length = 0
                                            column_letter = col[0].column_letter
                                            for cell in col:
                                                # 设置单元格自动换行并且垂直居中
                                                cell.alignment = Alignment(wrapText=True, vertical='center')
                                                
                                                try: 
                                                    # 汉字宽度大致是英文字符的 1.7 - 2 倍，我们粗略计算以防挤在一起
                                                    width_est = len(str(cell.value).encode('gbk'))
                                                    if width_est > max_length:
                                                        max_length = width_est
                                                except:
                                                    pass
                                            # 设置列宽为自适应，但最高不超过 60 避免一列霸占整个屏幕
                                            ws.column_dimensions[column_letter].width = min((max_length + 2), 60)
                                            
                                        print(f"✅ 已将提取到的纯文本工作记录精算填入 Excel 专属 Sheet: '{safe_sheet_name}'。")
                                    else:
                                        print("⚠️ 解析表格 DOM 失败，页面中未发现标准工作记录结构！")
                                        
                                except Exception as e:
                                    print(f"⚠️ 解析提取工作记录表格异常，跳过该项目: {e}")
                            else:
                                print("⚠️ 未能在弹窗中找到'工作记录'页签，可能不是预期的详情页结构。")
                                
                            # 3. 关闭侧滑/弹窗回到报表
                            print("关闭当前弹层并准备下一个拉取循环...")
                            # 关闭抽屉前先把鼠标移走，避免触发 Canvas 的残留事件
                            page.mouse.move(0, 0)
                            page.wait_for_timeout(200)
                            try:
                                close_btns = page.locator(".el-drawer__close-btn, .icon-close, .close-btn, button:has-text('关闭')").all()
                                for cbtn in reversed(close_btns): # 有时最新出来的 X 按钮在最后
                                    if cbtn.is_visible():
                                        cbtn.hover()
                                        cbtn.click(force=True)
                                        break
                                # 备选关闭法：按下主键盘 ESC
                                page.keyboard.press("Escape")
                            except: pass
                            page.wait_for_timeout(1500)  # 优化：2500ms→1500ms
                        else:
                            print(f"⚠️ 在当前可视报表中未能找到项目超链接：'{proj_name_str}'，它可能位于下一页或被隐藏。")
                    
                    print("\n💾 所有可用图像已插入完毕，正在封装最终富媒体 Excel 报告...")
                    wb.save(save_path)
                    print(f"🎉🎉 第二阶段：数据分析整合大圆满！全量报告固化至:\n {save_path}")
                    
                    # ==================== 第三阶段：PDF 导出 ====================
                    print("\n" + "=" * 50)
                    print("🚀 开始第三阶段：将 Excel 报告转换为 PDF！")
                    print("=" * 50)
                    try:
                        script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'convert_pdf.py')
                        result = subprocess.run(
                            [sys.executable, script_path, save_path],
                            capture_output=False,
                            text=True
                        )
                        if result.returncode != 0:
                            print(f"⚠️ PDF 导出进程异常退出 (code={result.returncode})")
                    except Exception as pdf_err:
                        print(f"⚠️ PDF 导出失败（不影响 Excel 报告）: {pdf_err}")
                    
                    # ==================== 第四阶段：Gemini AI 智能分析 ====================
                    print("\n" + "=" * 50)
                    print("🧠 开始第四阶段：Gemini AI 销售数据智能分析！")
                    print("=" * 50)
                    try:
                        analysis_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'analyze_sales.py')
                        result = subprocess.run(
                            [sys.executable, analysis_script, save_path],
                            capture_output=False,
                            text=True
                        )
                        if result.returncode != 0:
                            print(f"⚠️ AI 分析进程异常退出 (code={result.returncode})，不影响已生成的 Excel 和 PDF 报告。")
                    except Exception as ai_err:
                        print(f"⚠️ AI 分析失败（不影响 Excel/PDF 报告）: {ai_err}")
                    
                    # ==================== 第五阶段：分区销售 Review 报告 ====================
                    print("\n" + "=" * 50)
                    print("📍 开始第五阶段：分区销售 Review 报告生成！")
                    print("=" * 50)
                    try:
                        region_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'analyze_region.py')
                        result = subprocess.run(
                            [sys.executable, region_script, save_path],
                            capture_output=False,
                            text=True
                        )
                        if result.returncode != 0:
                            print(f"⚠️ 分区分析进程异常退出 (code={result.returncode})，不影响已生成的全局报告。")
                    except Exception as region_err:
                        print(f"⚠️ 分区分析失败（不影响全局报告）: {region_err}")
                    
                    # ==================== 第六阶段：单人销售工作报告 ====================
                    print("\n" + "=" * 50)
                    print("📍 开始第六阶段：单人销售线索分析报告生成！")
                    print("=" * 50)
                    try:
                        person_script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'analyze_person.py')
                        result = subprocess.run(
                            [sys.executable, person_script, save_path],
                            capture_output=False,
                            text=True
                        )
                        if result.returncode != 0:
                            print(f"⚠️ 单人分析进程异常退出 (code={result.returncode})，不影响已生成的全局/分区报告。")
                    except Exception as person_err:
                        print(f"⚠️ 单人分析失败（不影响全局/分区报告）: {person_err}")
            except Exception as ex:
                print(f"\n❌ 第二阶段 Excel 读取/融合操作中途发生异常: {ex}")
        except Exception as e:
            print(f"\n❌ 操作流严重中断，抓取失败: {e}")
            
        print("-" * 50)
        
        # ==================== 第七阶段：清理临时文件 ====================
        print("\n🧹 正在自动清理临时文件 (*.html, *.md, *.webm)...")
        import glob
        for ext in ['*.html', '*.md', '*.webm']:
            for f in glob.glob(os.path.join(DOWNLOAD_DIR, ext)):
                try:
                    os.remove(f)
                except:
                    pass
        print("✨ 临时报告底稿和视频记录清除完毕！")
        
        print("\n🎉 全部自动化任务已顺利完成，您可以直接查阅对应报告了！\n")
        
        page.wait_for_timeout(5000) 
        browser.close()

if __name__ == "__main__":
    run()
