#!/usr/bin/env python3
"""
Gemini 销售数据智能分析工具 (v2.0 - 2026-03-12 稳定版)

读取 Excel 结构化数据 → 调用 Gemini API 进行三维度分析 → 输出 PDF 报告

▸ 分析流程 (三轮 Gemini API 调用)：
  1. 销售机会可能性评估 — 逐项诊断 + 实战行动建议 + 管理层三封信
  2. 片区与个人绩效分析 — 片区对比 + 个人表现 + 改进建议
  3. 三区域百分制综合评分 — 量化评分矩阵 + 僵尸线索明细 + 排名总评

▸ AI Prompt 定义在: sales_ai_prompts.py (可独立修改，无需改动本文件)
▸ PDF 样式定义在: generate_report_pdf() 函数内 (HTML+CSS 模板)
▸ 片区人员映射: REGION_MAP 字典 (如有人员调整需同步修改)
"""

import os
import sys
import glob
import html
print("=== 调试: 脚本已启动 ===", flush=True)
import tempfile
from google import genai
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


# ==================== 配置 (已固化) ====================
from dotenv import load_dotenv
load_dotenv()
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-2.5-flash-lite"
   # 2026-03-12 验证通过的模型版本

# 配置代理以防止国内直连报 "User location is not supported"
os.environ["HTTP_PROXY"] = "http://127.0.0.1:7890"
os.environ["HTTPS_PROXY"] = "http://127.0.0.1:7890"


import json
def load_org_config():
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception:
        config = {}
    org_chart = config.get("org_chart", {})
    region_map = {}
    for region, data in org_chart.items():
        if region == "支持部门": continue
        people = [data["manager"]["name"]]
        for sub in data.get("subordinates", []):
            people.append(sub["name"])
        region_map[region] = people
    
    lines = ["## 组织架构、汇报关系与 2026 年度销售目标"]
    for region, data in org_chart.items():
        mgr = data["manager"]
        postfix = "" if region == "支持部门" else "片区"
        
        region_quota_str = f" [片区总任务: {data['quota']}万]" if "quota" in data else ""
        mgr_quota_str = f" [个人年度任务: {mgr['quota']}万]" if "quota" in mgr else ""
        lines.append(f"- **{region}{postfix}{mgr['title']}**：{mgr['name']}{region_quota_str}{mgr_quota_str}")
        for sub in data.get("subordinates", []):
            sub_quota_str = f" [个人年度任务: {sub['quota']}万]" if "quota" in sub else ""
            lines.append(f"  - 下属{sub['title']}：{sub['name']}{sub_quota_str}")
    org_chart_md = "\n".join(lines)
    return region_map, org_chart_md

REGION_MAP, ORG_CHART_MD = load_org_config()

# 反向映射：人名 → 片区
PERSON_TO_REGION = {}
for region, people in REGION_MAP.items():
    for person in people:
        PERSON_TO_REGION[person] = region


# ==================== 数据读取 ====================
def read_excel_data(excel_path):
    """读取 Excel 并返回结构化数据"""
    wb = load_workbook(excel_path)
    
    # 1. 读取主报表
    main_sheet = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in main_sheet[1]]
    
    main_data = []
    safe_to_original = {}
    import re
    for row in main_sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        main_data.append(record)
        # 建立 sheet name 映射表
        orig_name = str(record.get('项目名称', '')).strip()
        safe_name = re.sub(r'[\\\\*?:/\\[\\]]', '_', orig_name)[:31]
        safe_to_original[safe_name] = orig_name
    
    # 2. 读取工作记录 sheets
    work_records = {}
    for sheet_name in wb.sheetnames[2:]:  # 跳过前两个 sheet（主报表 + 元数据）
        ws = wb[sheet_name]
        if ws.max_row and ws.max_row > 1:
            ws_headers = [cell.value for cell in ws[1]]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(ws_headers, row)))
            # 还原为原始项目名称，保证 AI 匹配一致性
            orig_name = safe_to_original.get(sheet_name, sheet_name)
            work_records[orig_name] = records
            
    return headers, main_data, work_records


def format_main_table(headers, main_data):
    """将主报表数据转为 Markdown 表格"""
    # 选择关键列
    key_cols = ['客\n户名称', '项目名称', '负责人', '商机阶段', '预期阶段', 
                '项目金额', '预计开票日期', '预计开票金额']
    
    # 找到实际存在的列
    available_cols = [c for c in key_cols if c in headers]
    if not available_cols:
        available_cols = headers[:8]
    
    # 清理列名中的换行符
    clean_cols = [c.replace('\n', '') for c in available_cols]
    
    lines = []
    lines.append('| ' + ' | '.join(clean_cols) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(available_cols)) + ' |')
    
    for record in main_data:
        vals = []
        for col in available_cols:
            v = record.get(col, '--')
            v = str(v).replace('\n', ' ').strip() if v else '--'
            vals.append(v)
        lines.append('| ' + ' | '.join(vals) + ' |')
    
    return '\n'.join(lines)


def format_work_records(work_records):
    """将工作记录转为结构化文本"""
    lines = []
    for project, records in work_records.items():
        lines.append(f"\n### {project}")
        for r in records:
            content = str(r.get('记录内容', '--')).replace('\n', ' ')
            follow_type = str(r.get('跟进类型', '--'))
            date = str(r.get('日计划', '--'))
            lines.append(f"- [{date}] ({follow_type}) {content}")
    return '\n'.join(lines)


# ==================== Prompt 导入 ====================
from sales_ai_prompts import build_prompt_1, build_prompt_2, build_prompt_3

def get_region_text(main_data, REGION_MAP):
    region_summary = {}
    for region, people in REGION_MAP.items():
        region_projects = []
        for record in main_data:
            responsible = str(record.get("负责人", ""))
            for person in people:
                if person in responsible:
                    region_projects.append(record)
                    break
        region_summary[region] = region_projects
    region_text = ""
    for region, people in REGION_MAP.items():
        joined_people = ", ".join(people)
        region_text += f"\n### {region}片区（{joined_people}）\n"
        projects = region_summary.get(region, [])
        for p in projects:
            name = str(p.get("项目名称", "--")).replace("\n", "")
            responsible = str(p.get("负责人", "--"))
            stage = str(p.get("商机阶段", "--"))
            amount = str(p.get("预计开票金额", "--"))
            date = str(p.get("预计开票日期", "--")).replace("\n", " ")
            region_text += f"- {name} | 负责人:{responsible} | 阶段:{stage} | 预计金额:{amount} | 开票日期:{date}\n"
    return region_text

# ==================== Gemini API 调用 ====================
def call_gemini(prompt, task_name):
    """调用 Gemini API（新版 google-genai SDK，带自动重试）"""
    import time as _time
    
    client = genai.Client(
        api_key=GEMINI_API_KEY,
        http_options={'api_version': 'v1'}
    )
    
    for attempt in range(3):
        try:
            print(f"\n🤖 正在请求 Gemini ({GEMINI_MODEL}) 分析: {task_name} ...", flush=True)
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=prompt
            )
            result = response.text
            print(f"   ✅ {task_name} 分析完成 ({len(result)} 字)")
            return result
        except Exception as e:
            err_str = str(e)
            print(f"   ❌ API 调用失败: {err_str}")
            if any(k in err_str for k in ['429', 'RESOURCE_EXHAUSTED', 'quota', '503', 'UNAVAILABLE', 'high demand', 'SSL', 'EOF', 'ConnectError', 'Timeout', 'Connection', 'Proxy', 'RemoteProtocolError', 'disconnected']):
                wait = 60 * (attempt + 1)
                print(f"   ⏳ 触发自动重试机制，等待 {wait} 秒后重试 ({attempt+1}/3)...")
                _time.sleep(wait)
            else:
                raise
    
    raise Exception(f"Gemini API 多次重试后仍然失败，请检查 API Key 的配额设置。\n"
                    f"当前配额层级为免费层，建议到 https://aistudio.google.com/ 点击'设置结算信息'升级为付费计划。")


# ==================== PDF 报告生成 ====================
def generate_report_pdf(analysis_1, analysis_2, analysis_3, pdf_path):
    """将三段分析结果渲染为 PDF"""
    
    import re
    from datetime import datetime
    
    today = datetime.now().strftime("%Y年%m月%d日")
    
    import markdown
    
    def md_to_html(md_text):
        """使用 markdown 库转换为 HTML，自动修复无空行或带缩进的表格"""
        lines = md_text.split('\\n')
        fixed_lines = []
        in_table = False
        for line in lines:
            stripped = line.strip()
            is_table_row = stripped.startswith('|') and stripped.count('|') >= 2
            
            if is_table_row:
                if not in_table:
                    if fixed_lines and fixed_lines[-1].strip() != '':
                        fixed_lines.append('')
                    in_table = True
                fixed_lines.append(line.lstrip(' \\t'))
            else:
                in_table = False
                fixed_lines.append(line)
                
        return markdown.markdown('\\n'.join(fixed_lines), extensions=['tables'])
    
    report_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
    @page {{ margin: 20mm 15mm; }}
    * {{ box-sizing: border-box; }}
    body {{
        font-family: "PingFang SC", "Microsoft YaHei", "Hiragino Sans GB", sans-serif;
        color: #1a1a1a;
        line-height: 1.7;
        margin: 0;
        padding: 0;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }}
    
    /* 封面 */
    .cover {{
        text-align: center;
        padding-top: 200px;
        page-break-after: always;
    }}
    .cover h1 {{
        font-size: 32px;
        color: #1e40af;
        margin-bottom: 20px;
    }}
    .cover .subtitle {{
        font-size: 16px;
        color: #64748b;
        margin: 8px 0;
    }}
    .cover .divider {{
        width: 120px;
        height: 3px;
        background: linear-gradient(90deg, #3b82f6, #8b5cf6);
        margin: 30px auto;
    }}
    
    /* 章节 */
    .chapter {{
        page-break-before: always;
    }}
    .chapter-header {{
        background: linear-gradient(135deg, #1e40af, #3b82f6);
        color: white;
        padding: 16px 24px;
        border-radius: 8px;
        margin-bottom: 20px;
    }}
    .chapter-header h1 {{
        font-size: 22px;
        margin: 0;
    }}
    .chapter-header p {{
        font-size: 13px;
        opacity: 0.85;
        margin: 4px 0 0;
    }}
    
    /* 表格 */
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 9px;
        margin: 12px 0;
        table-layout: auto;
    }}
    th {{
        background: #1e40af;
        color: white;
        padding: 4px 6px;
        text-align: center;
        font-size: 9px;
        border: 1px solid #1e40af;
    }}
    td {{
        border: 1px solid #e2e8f0;
        padding: 4px 6px;
        vertical-align: top;
        word-break: break-all;
        word-wrap: break-word;
    }}
    tr {{
        page-break-inside: avoid;
    }}
    tr:nth-child(even) {{ background: #f1f5f9; }}
    
    h2 {{
        color: #1e40af;
        font-size: 16px;
        border-left: 4px solid #3b82f6;
        padding-left: 12px;
        margin-top: 24px;
    }}
    h3 {{
        color: #334155;
        font-size: 14px;
        margin-top: 16px;
    }}
    
    p {{ font-size: 12px; margin: 6px 0; }}
    ul {{ font-size: 12px; padding-left: 20px; }}
    li {{ margin: 4px 0; }}
    strong {{ color: #1e40af; }}
    
    .footer {{
        text-align: center;
        font-size: 9px;
        color: #94a3b8;
        margin-top: 40px;
        padding-top: 12px;
        border-top: 1px solid #e2e8f0;
    }}
</style>
</head>
<body>

<!-- 封面 -->
<div class="cover">
    <h1>📊 销售机会智能分析报告</h1>
    <div class="divider"></div>
    <p class="subtitle">2026年大额开票机会统计 · 深度分析</p>
    <p class="subtitle">{today}</p>
    <p class="subtitle" style="margin-top: 40px; font-size: 12px; color: #94a3b8;">
        Powered by Gemini AI · 自动化商务智能分析
    </p>
</div>

<!-- 第一章 -->
<div class="chapter">
    <div class="chapter-header">
        <h1>第一章 · 销售机会可能性评估</h1>
        <p>基于机会阶段、金额体量和跟进记录，评估每个机会的成单可能性</p>
    </div>
    {md_to_html(analysis_1)}
</div>

<!-- 第二章 -->
<div class="chapter">
    <div class="chapter-header">
        <h1>第二章 · 片区与个人绩效分析</h1>
        <p>华东 / 华南 / 华北三大片区的商务能力对比与个人表现评估</p>
    </div>
    {md_to_html(analysis_2)}
</div>

<!-- 第三章 -->
<div class="chapter">
    <div class="chapter-header">
        <h1>第三章 · 三区域百分制综合评分</h1>
        <p>从 Leads 规模、线索真实度、跟进质量三维度进行量化评分</p>
    </div>
    {md_to_html(analysis_3)}
</div>

<div class="footer">
    本报告由 Gemini AI 自动生成 · {today} · 仅供内部参考
</div>

</body>
</html>"""
    # 写入临时 HTML
    with open('debug_report.html', 'w', encoding='utf-8') as f:
        f.write(report_html)
    tmp = tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8')
    tmp.write(report_html)
    tmp.close()
    
    # Playwright 导出 PDF
    print(f"\n🖨️ 正在渲染分析报告 PDF ...")
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f'file://{tmp.name}', wait_until='networkidle')
        page.wait_for_timeout(1000)
        
        page.pdf(
            path=pdf_path,
            format='A4',
            print_background=True,
            margin={'top': '15mm', 'bottom': '15mm', 'left': '12mm', 'right': '12mm'}
        )
        browser.close()
    
    os.unlink(tmp.name)
    
    file_size = os.path.getsize(pdf_path)
    size_str = f"{file_size/1024:.0f}KB" if file_size < 1024*1024 else f"{file_size/1024/1024:.1f}MB"
    print(f"\n🎉🎉 销售分析报告 PDF 生成成功！")
    print(f"   📁 路径: {pdf_path}")
    print(f"   📏 大小: {size_str}")


# ==================== 主流程 ====================
def analyze_sales(excel_path, pdf_path=None):
    """主入口：读取 Excel → Gemini 分析 → 生成 PDF"""
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return
    
    if pdf_path is None:
        # 获取日期后缀，从文件名中提取或使用当前日期
        base = os.path.basename(excel_path)
        # 尝试从文件名中提取日期 (原格式类似 HAO_2026年大额开票机会统计_20260311_180530.xlsx)
        import re
        date_match = re.search(r'_(\d{8})_', base)
        if date_match:
            date_str = date_match.group(1)
        else:
            from datetime import datetime
            date_str = datetime.now().strftime("%Y%m%d")
            
        dir_name = os.path.dirname(excel_path)
        pdf_name = f"益高重点项目+SalesAiReviewReport（{date_str}）.pdf"
        pdf_path = os.path.join(dir_name, pdf_name)
    
    print("=" * 60)
    print("🧠 Gemini 销售智能分析系统 启动")
    print("=" * 60)
    
    # 1. 读取数据
    print(f"\n📖 读取 Excel 数据: {os.path.basename(excel_path)}")
    headers, main_data, work_records = read_excel_data(excel_path)
    print(f"   ✅ 主报表: {len(main_data)} 条记录, {len(headers)} 列")
    print(f"   ✅ 工作记录: {len(work_records)} 个项目")
    
    # 2. 格式化数据
    main_table_md = format_main_table(headers, main_data)
    work_records_md = format_work_records(work_records)
    
    # 3. 构建并执行三轮分析
    prompt_1 = build_prompt_1(main_table_md, work_records_md, ORG_CHART_MD)
    analysis_1 = call_gemini(prompt_1, "销售机会可能性评估")
    
    region_text = get_region_text(main_data, REGION_MAP)
    prompt_2 = build_prompt_2(region_text, main_table_md, work_records_md, REGION_MAP, ORG_CHART_MD)
    analysis_2 = call_gemini(prompt_2, "片区与个人绩效分析")
    
    prompt_3 = build_prompt_3(main_table_md, work_records_md, ORG_CHART_MD)
    analysis_3 = call_gemini(prompt_3, "三区域百分制综合评分")
    

    # 4. 生成 PDF 报告
    generate_report_pdf(analysis_1, analysis_2, analysis_3, pdf_path)



def find_latest_excel():
    """自动找到 downloads 目录下最新的 Excel 文件"""
    download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')
    xlsx_files = [f for f in glob.glob(os.path.join(download_dir, '**', '*.xlsx'), recursive=True) 
                  if not os.path.basename(f).startswith('~$')]
    if not xlsx_files:
        print("❌ downloads 目录下没有找到 .xlsx 文件")
        return None
    return max(xlsx_files, key=os.path.getmtime)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_excel()
    
    if excel_file:
        analyze_sales(excel_file)
