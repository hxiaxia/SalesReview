#!/usr/bin/env python3
"""
分区销售 Review 分析工具 (v1.0 - 2026-03-12)

按地区拆分 Excel 数据，为每个地区生成独立的 AI 分析 PDF 报告。
▸ 多人片区（华东/华北）：个人间横向对比与排名
▸ 单人片区（华南）：个人对标行业基准自查评分

本脚本独立于第 4 步的 analyze_sales.py，不修改任何前序代码。
"""

import os
import sys
import glob
import html
import tempfile
from google import genai
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from region_ai_prompts import (
    build_region_prompt_1,
    build_region_prompt_2,
    build_region_prompt_2_solo,
    build_region_prompt_3,
    build_region_prompt_3_solo,
)

# ==================== 配置 (与 analyze_sales.py 保持一致) ====================
from dotenv import load_dotenv
load_dotenv()
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = "gemini-2.5-flash-lite"

os.environ["HTTP_PROXY"] = "http://127.0.0.1:7890"
os.environ["HTTPS_PROXY"] = "http://127.0.0.1:7890"

import json
def load_org_config():
    config_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "config.json")
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config = json.load(f)
    except Exception:
        config = {}
    org_chart = config.get("org_chart", {})
    region_map = {}
    for region, data in org_chart.items():
        if region == "支持部门": continue
        people = [data["manager"]["name"]]
        for sub in data.get("subordinates", []):
            people.append(sub["name"])
        region_map[region] = people
    
    lines = ["## 组织架构、汇报关系与 2026 年度销售目标"]
    for region, data in org_chart.items():
        mgr = data["manager"]
        postfix = "" if region == "支持部门" else "片区"
        
        region_quota_str = f" [片区总任务: {data['quota']}万]" if "quota" in data else ""
        mgr_quota_str = f" [个人年度任务: {mgr['quota']}万]" if "quota" in mgr else ""
        lines.append(f"- **{region}{postfix}{mgr['title']}**：{mgr['name']}{region_quota_str}{mgr_quota_str}")
        for sub in data.get("subordinates", []):
            sub_quota_str = f" [个人年度任务: {sub['quota']}万]" if "quota" in sub else ""
            lines.append(f"  - 下属{sub['title']}：{sub['name']}{sub_quota_str}")
    org_chart_md = "\n".join(lines)
    return region_map, org_chart_md

REGION_MAP, ORG_CHART_MD = load_org_config()


# ==================== 数据读取与筛选 ====================
def read_excel_data(excel_path):
    """读取 Excel 并返回结构化数据"""
    wb = load_workbook(excel_path)
    
    main_sheet = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in main_sheet[1]]
    
    main_data = []
    safe_to_original = {}
    import re
    for row in main_sheet.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        main_data.append(record)
        # 建立 sheet name 映射表
        orig_name = str(record.get('项目名称', '')).strip()
        safe_name = re.sub(r'[\\\\*?:/\\[\\]]', '_', orig_name)[:31]
        safe_to_original[safe_name] = orig_name
    
    work_records = {}
    for sheet_name in wb.sheetnames[2:]:
        ws = wb[sheet_name]
        if ws.max_row and ws.max_row > 1:
            ws_headers = [cell.value for cell in ws[1]]
            records = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                records.append(dict(zip(ws_headers, row)))
            # 还原为原始项目名称，保证 AI 匹配一致性
            orig_name = safe_to_original.get(sheet_name, sheet_name)
            work_records[orig_name] = records
            
    return headers, main_data, work_records


def filter_data_by_region(headers, main_data, work_records, people_list):
    """按片区人员筛选数据"""
    # 筛选主报表
    region_data = []
    region_project_names = set()
    for record in main_data:
        responsible = str(record.get("负责人", ""))
        for person in people_list:
            if person in responsible:
                region_data.append(record)
                proj_name = str(record.get("项目名称", "")).strip()
                if proj_name:
                    region_project_names.add(proj_name)
                break
    
    # 筛选工作记录：只保留本地区项目的工作记录 Sheet
    region_work_records = {}
    for sheet_name, records in work_records.items():
        # Sheet 名称经过截断(最长31字符)和特殊字符替换，需模糊匹配
        clean_sheet = sheet_name.strip()
        matched = False
        for proj_name in region_project_names:
            # 双向前缀匹配（因为 Sheet 名可能被截断）
            if clean_sheet in proj_name or proj_name[:28] in clean_sheet:
                matched = True
                break
        if matched:
            region_work_records[sheet_name] = records
    
    return region_data, region_work_records


# ==================== 格式化 ====================
def format_main_table(headers, data):
    """将数据转为 Markdown 表格"""
    key_cols = ['客\n户名称', '项目名称', '负责人', '商机阶段', '预期阶段',
                '项目金额', '预计开票日期', '预计开票金额']
    available_cols = [c for c in key_cols if c in headers]
    if not available_cols:
        available_cols = headers[:8]
    
    clean_cols = [c.replace('\n', '') for c in available_cols]
    lines = []
    lines.append('| ' + ' | '.join(clean_cols) + ' |')
    lines.append('| ' + ' | '.join(['---'] * len(available_cols)) + ' |')
    
    for record in data:
        vals = []
        for col in available_cols:
            v = record.get(col, '--')
            v = str(v).replace('\n', ' ').strip() if v else '--'
            vals.append(v)
        lines.append('| ' + ' | '.join(vals) + ' |')
    
    return '\n'.join(lines)


def format_work_records(work_records):
    """将工作记录转为结构化文本"""
    lines = []
    for project, records in work_records.items():
        lines.append(f"\n### {project}")
        for r in records:
            content = str(r.get('记录内容', '--')).replace('\n', ' ')
            follow_type = str(r.get('跟进类型', '--'))
            date = str(r.get('日计划', '--'))
            lines.append(f"- [{date}] ({follow_type}) {content}")
    return '\n'.join(lines)


# ==================== Gemini API 调用 ====================
def call_gemini(prompt, task_name):
    """调用 Gemini API（带自动重试）"""
    import time
    for attempt in range(3):
        try:
            client = genai.Client(api_key=GEMINI_API_KEY, http_options={'api_version': 'v1', 'timeout': 600000})
            print(f"\n🤖 正在请求 Gemini ({GEMINI_MODEL}) 分析: {task_name} ...", flush=True)
            response = client.models.generate_content(
                model=GEMINI_MODEL,
                contents=prompt
            )
            result = response.text
            print(f"   ✅ {task_name} 分析完成 ({len(result)} 字)", flush=True)
            return result
        except Exception as e:
            err_str = str(e)
            print(f"   ❌ API 调用失败: {err_str}", flush=True)
            if any(k.lower() in err_str.lower() for k in ['429', 'RESOURCE_EXHAUSTED', 'quota', '503', 'UNAVAILABLE', 'high demand', 'SSL', 'EOF', 'ConnectError', 'Timeout', 'Connection', 'Proxy', 'RemoteProtocolError', 'disconnected', 'timed out']):
                wait = 60 * (attempt + 1)
                print(f"   ⏳ 触发自动重试机制，等待 {wait} 秒后重试 ({attempt+1}/3)...", flush=True)
                time.sleep(wait)
            else:
                raise
    
    raise Exception(f"Gemini API 多次重试后仍然失败")


# ==================== PDF 报告生成 ====================
def generate_region_report_pdf(region_name, is_solo, analysis_1, analysis_2, analysis_3, pdf_path):
    """将三段分析结果渲染为地区 PDF 报告"""
    from datetime import datetime
    import markdown
    
    today = datetime.now().strftime("%Y年%m月%d日")
    
    def md_to_html(md_text):
        """使用 markdown 库转换为 HTML，自动修复无空行或带缩进的表格"""
        lines = md_text.split('\\n')
        fixed_lines = []
        in_table = False
        for line in lines:
            stripped = line.strip()
            is_table_row = stripped.startswith('|') and stripped.count('|') >= 2
            
            if is_table_row:
                if not in_table:
                    if fixed_lines and fixed_lines[-1].strip() != '':
                        fixed_lines.append('')
                    in_table = True
                fixed_lines.append(line.lstrip(' \\t'))
            else:
                in_table = False
                fixed_lines.append(line)
                
        return markdown.markdown('\\n'.join(fixed_lines), extensions=['tables'])
    
    # 根据模式调整章节标题
    if is_solo:
        ch2_title = "个人能力画像与行业对标"
        ch2_desc = "以行业优秀销售的基准进行自查对标分析"
        ch3_title = "个人对标行业基准评分"
        ch3_desc = "以行业标杆为满分参照系的绝对评分与差距分析"
    else:
        ch2_title = "个人表现对比分析"
        ch2_desc = f"{region_name}片区内各位销售的表现对比与改进建议"
        ch3_title = "个人百分制综合评分"
        ch3_desc = f"从 Leads 规模、线索真实度、跟进质量三维度对{region_name}片区每位销售量化评分"
    
    report_html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
    @page {{ margin: 20mm 15mm; }}
    * {{ box-sizing: border-box; }}
    body {{
        font-family: "PingFang SC", "Microsoft YaHei", "Hiragino Sans GB", sans-serif;
        color: #1a1a1a;
        line-height: 1.7;
        margin: 0;
        padding: 0;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }}
    
    .cover {{
        text-align: center;
        padding-top: 200px;
        page-break-after: always;
    }}
    .cover h1 {{
        font-size: 32px;
        color: #1e40af;
        margin-bottom: 20px;
    }}
    .cover .subtitle {{
        font-size: 16px;
        color: #64748b;
        margin: 8px 0;
    }}
    .cover .divider {{
        width: 120px;
        height: 3px;
        background: linear-gradient(90deg, #3b82f6, #8b5cf6);
        margin: 30px auto;
    }}
    
    .chapter {{
        page-break-before: always;
    }}
    .chapter-header {{
        background: linear-gradient(135deg, #1e40af, #3b82f6);
        color: white;
        padding: 16px 24px;
        border-radius: 8px;
        margin-bottom: 20px;
    }}
    .chapter-header h1 {{
        font-size: 22px;
        margin: 0;
    }}
    .chapter-header p {{
        font-size: 13px;
        opacity: 0.85;
        margin: 4px 0 0;
    }}
    
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 9px;
        margin: 12px 0;
        table-layout: auto;
    }}
    th {{
        background: #1e40af;
        color: white;
        padding: 4px 6px;
        text-align: center;
        font-size: 9px;
        border: 1px solid #1e40af;
    }}
    td {{
        border: 1px solid #e2e8f0;
        padding: 4px 6px;
        vertical-align: top;
        word-break: break-all;
        word-wrap: break-word;
    }}
    tr {{
        page-break-inside: avoid;
    }}
    tr:nth-child(even) {{ background: #f1f5f9; }}
    
    h2 {{
        color: #1e40af;
        font-size: 16px;
        border-left: 4px solid #3b82f6;
        padding-left: 12px;
        margin-top: 24px;
    }}
    h3 {{
        color: #334155;
        font-size: 14px;
        margin-top: 16px;
    }}
    
    p {{ font-size: 12px; margin: 6px 0; }}
    ul {{ font-size: 12px; padding-left: 20px; }}
    li {{ margin: 4px 0; }}
    strong {{ color: #1e40af; }}
    
    .footer {{
        text-align: center;
        font-size: 9px;
        color: #94a3b8;
        margin-top: 40px;
        padding-top: 12px;
        border-top: 1px solid #e2e8f0;
    }}
</style>
</head>
<body>

<div class="cover">
    <h1>📊 {region_name}片区 销售机会分析报告</h1>
    <div class="divider"></div>
    <p class="subtitle">2026年大额开票机会统计 · {region_name}片区深度分析</p>
    <p class="subtitle">{today}</p>
    <p class="subtitle" style="margin-top: 40px; font-size: 12px; color: #94a3b8;">
        Powered by Gemini AI · 自动化商务智能分析
    </p>
</div>

<div class="chapter">
    <div class="chapter-header">
        <h1>第一章 · 销售机会可能性评估</h1>
        <p>{region_name}片区机会阶段、金额体量和跟进记录的深度诊断</p>
    </div>
    {md_to_html(analysis_1)}
</div>

<div class="chapter">
    <div class="chapter-header">
        <h1>第二章 · {ch2_title}</h1>
        <p>{ch2_desc}</p>
    </div>
    {md_to_html(analysis_2)}
</div>

<div class="chapter">
    <div class="chapter-header">
        <h1>第三章 · {ch3_title}</h1>
        <p>{ch3_desc}</p>
    </div>
    {md_to_html(analysis_3)}
</div>

<div class="footer">
    本报告由 Gemini AI 自动生成 · {today} · {region_name}片区 · 仅供内部参考
</div>

</body>
</html>"""
    
    tmp = tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8')
    tmp.write(report_html)
    tmp.close()
    
    print(f"\n🖨️ 正在渲染 {region_name}片区 分析报告 PDF ...", flush=True)
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(f'file://{tmp.name}', wait_until='networkidle')
        page.wait_for_timeout(1000)
        
        page.pdf(
            path=pdf_path,
            format='A4',
            print_background=True,
            margin={'top': '15mm', 'bottom': '15mm', 'left': '12mm', 'right': '12mm'}
        )
        browser.close()
    
    os.unlink(tmp.name)
    
    file_size = os.path.getsize(pdf_path)
    size_str = f"{file_size/1024:.0f}KB" if file_size < 1024*1024 else f"{file_size/1024/1024:.1f}MB"
    print(f"\n🎉 {region_name}片区分析报告 PDF 生成成功！", flush=True)
    print(f"   📁 路径: {pdf_path}", flush=True)
    print(f"   📏 大小: {size_str}", flush=True)


# ==================== 主流程 ====================
def analyze_regions(excel_path):
    """主入口：读取 Excel → 按地区拆分 → 逐地区 Gemini 分析 → 生成各地区 PDF"""
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}", flush=True)
        return
    
    print("=" * 60, flush=True)
    print("🧠 分区销售 Review 分析系统 启动", flush=True)
    print("=" * 60, flush=True)
    
    # 1. 读取全量数据
    print(f"\n📖 读取 Excel 数据: {os.path.basename(excel_path)}", flush=True)
    headers, main_data, work_records = read_excel_data(excel_path)
    print(f"   ✅ 主报表: {len(main_data)} 条记录, {len(headers)} 列", flush=True)
    print(f"   ✅ 工作记录: {len(work_records)} 个项目", flush=True)
    
    base_name = os.path.splitext(excel_path)[0]
    
    # 2. 逐地区处理
    for region_name, people_list in list(REGION_MAP.items()): # 遍历所有片区
        print(f"\n{'='*60}", flush=True)
        print(f"📍 开始处理: {region_name}片区 ({', '.join(people_list)})", flush=True)
        print(f"{'='*60}", flush=True)
        
        is_solo = len(people_list) == 1
        mode_name = "单人自查模式" if is_solo else f"多人对比模式 ({len(people_list)}人)"
        print(f"   📋 分析模式: {mode_name}", flush=True)
        
        # 筛选数据
        region_data, region_work_records = filter_data_by_region(
            headers, main_data, work_records, people_list
        )
        print(f"   ✅ 筛选完成: {len(region_data)} 条记录, {len(region_work_records)} 个工作记录", flush=True)
        
        if len(region_data) == 0:
            print(f"   ⚠️ {region_name}片区无数据，跳过。", flush=True)
            continue
        
        # 格式化
        region_table_md = format_main_table(headers, region_data)
        region_records_md = format_work_records(region_work_records)
        
        # 三轮 Gemini 分析
        # Prompt 1：销售机会评估（通用）
        prompt_1 = build_region_prompt_1(region_name, people_list, region_table_md, region_records_md, ORG_CHART_MD)
        analysis_1 = call_gemini(prompt_1, f"{region_name}片区 - 销售机会评估")
        
        # Prompt 2：个人对比 / 行业对标
        if is_solo:
            prompt_2 = build_region_prompt_2_solo(region_name, people_list[0], region_table_md, region_records_md, ORG_CHART_MD)
        else:
            prompt_2 = build_region_prompt_2(region_name, people_list, region_table_md, region_records_md, ORG_CHART_MD)
        analysis_2 = call_gemini(prompt_2, f"{region_name}片区 - {'行业对标' if is_solo else '个人对比'}")
        
        # Prompt 3：个人评分 / 绝对评分
        if is_solo:
            prompt_3 = build_region_prompt_3_solo(region_name, people_list[0], region_table_md, region_records_md, ORG_CHART_MD)
        else:
            prompt_3 = build_region_prompt_3(region_name, people_list, region_table_md, region_records_md, ORG_CHART_MD)
        analysis_3 = call_gemini(prompt_3, f"{region_name}片区 - {'绝对评分' if is_solo else '个人评分'}")
        

        # 生成 PDF
        base_filename = os.path.basename(excel_path)
        import re
        date_match = re.search(r'_(\d{8})_', base_filename)
        if date_match:
            date_str = date_match.group(1)
        else:
            from datetime import datetime
            date_str = datetime.now().strftime("%Y%m%d")

        pdf_name = f"{region_name}重点项目+SalesAiReviewReport（{date_str}）.pdf"
        dir_name = os.path.dirname(excel_path)
        region_dir = os.path.join(dir_name, region_name)
        os.makedirs(region_dir, exist_ok=True)
        pdf_path = os.path.join(region_dir, pdf_name)
        generate_region_report_pdf(region_name, is_solo, analysis_1, analysis_2, analysis_3, pdf_path)
    
    print(f"\n{'='*60}", flush=True)
    print(f"🎉🎉 全部地区分析完成！", flush=True)
    print(f"{'='*60}", flush=True)


def find_latest_excel():
    """自动找到 downloads 目录下最新的 Excel 文件"""
    download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')
    xlsx_files = [f for f in glob.glob(os.path.join(download_dir, '**', '*.xlsx'), recursive=True)
                  if not os.path.basename(f).startswith('~$')]
    if not xlsx_files:
        print("❌ downloads 目录下没有找到 .xlsx 文件", flush=True)
        return None
    return max(xlsx_files, key=os.path.getmtime)


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_excel()
    
    if excel_file:
        analyze_regions(excel_file)
