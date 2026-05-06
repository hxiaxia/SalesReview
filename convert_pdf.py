#!/usr/bin/env python3
"""
Excel → PDF 转换工具
利用 openpyxl 读取 Excel 数据，生成带样式的 HTML，再通过 Playwright 无头浏览器导出 PDF。
无需安装任何额外依赖。
"""

import os
import sys
import glob
import html
import tempfile
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


def excel_to_pdf(excel_path: str, pdf_path: str = None):
    """将 Excel 文件的所有 Sheet 转换为一个 PDF 文件"""
    
    if not os.path.exists(excel_path):
        print(f"❌ 文件不存在: {excel_path}")
        return
    
    if pdf_path is None:
        pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
    
    print(f"📖 正在读取 Excel: {os.path.basename(excel_path)}")
    wb = load_workbook(excel_path)
    
    # ==================== 生成 HTML ====================
    html_parts = []
    
    # HTML 头部 + 全局样式
    html_parts.append("""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<style>
    @page { margin: 15mm 10mm; }
    
    * { box-sizing: border-box; }
    
    body {
        font-family: "PingFang SC", "Microsoft YaHei", "Hiragino Sans GB", "Noto Sans CJK SC", sans-serif;
        color: #1a1a1a;
        margin: 0;
        padding: 0;
        -webkit-print-color-adjust: exact;
        print-color-adjust: exact;
    }
    
    .sheet-section {
        page-break-after: always;
    }
    .sheet-section:last-child {
        page-break-after: avoid;
    }
    
    .sheet-title {
        font-size: 16px;
        font-weight: 700;
        color: #1a56db;
        border-bottom: 2px solid #1a56db;
        padding-bottom: 6px;
        margin-bottom: 12px;
    }
    
    table {
        width: 100%;
        border-collapse: collapse;
        font-size: 11px;
        table-layout: fixed;
    }
    
    th, td {
        border: 1px solid #cbd5e1;
        padding: 5px 6px;
        text-align: left;
        vertical-align: top;
        word-wrap: break-word;
        overflow-wrap: break-word;
    }
    
    th {
        background-color: #1e40af;
        color: white;
        font-weight: 600;
        font-size: 10px;
        text-align: center;
        white-space: nowrap;
    }
    
    tr:nth-child(even) {
        background-color: #f1f5f9;
    }
    
    tr:hover {
        background-color: #e2e8f0;
    }
    
    /* 主报表特殊样式（15列宽表） */
    .wide-table {
        font-size: 8px;
    }
    .wide-table th {
        font-size: 7.5px;
        padding: 3px 2px;
    }
    .wide-table td {
        padding: 3px 2px;
    }
    
    /* 元数据 sheet */
    .meta-content {
        white-space: pre-wrap;
        font-size: 11px;
        line-height: 1.6;
        color: #475569;
        background: #f8fafc;
        padding: 12px;
        border-radius: 6px;
        border: 1px solid #e2e8f0;
    }
    
    /* 工作记录表第一列（记录内容）加宽 */
    .work-record-table td:first-child {
        width: 40%;
    }
    
    .page-footer {
        text-align: center;
        font-size: 9px;
        color: #94a3b8;
        margin-top: 8px;
    }
</style>
</head>
<body>
""")
    
    total_sheets = len(wb.sheetnames)
    
    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row is None or max_row == 0:
            continue
        
        print(f"   📄 渲染 Sheet [{idx+1}/{total_sheets}]: '{sheet_name}' ({max_row}行 × {max_col}列)")
        
        html_parts.append(f'<div class="sheet-section">')
        html_parts.append(f'<div class="sheet-title">📋 {html.escape(sheet_name)}</div>')
        
        # 判断 Sheet 类型
        is_main_report = (max_col >= 10)  # 主报表15列
        is_meta = (sheet_name == '数据范围和系统提示')
        
        if is_meta:
            # 元数据 Sheet：直接以文本块形式呈现
            for row in ws.iter_rows(values_only=True):
                for cell_val in row:
                    if cell_val:
                        html_parts.append(f'<div class="meta-content">{html.escape(str(cell_val))}</div>')
        else:
            # 表格 Sheet
            table_class = 'wide-table' if is_main_report else 'work-record-table'
            html_parts.append(f'<table class="{table_class}">')
            
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx == 0:
                    # 表头行
                    html_parts.append('<thead><tr>')
                    for cell_val in row:
                        val = html.escape(str(cell_val).strip()) if cell_val else ''
                        html_parts.append(f'<th>{val}</th>')
                    html_parts.append('</tr></thead><tbody>')
                else:
                    # 数据行
                    html_parts.append('<tr>')
                    for cell_val in row:
                        val = html.escape(str(cell_val).strip()) if cell_val else '--'
                        html_parts.append(f'<td>{val}</td>')
                    html_parts.append('</tr>')
            
            html_parts.append('</tbody></table>')
        
        html_parts.append(f'<div class="page-footer">第 {idx+1}/{total_sheets} 页 · {html.escape(sheet_name)}</div>')
        html_parts.append('</div>')  # end sheet-section
    
    html_parts.append('</body></html>')
    html_content = '\n'.join(html_parts)
    
    # ==================== 写入临时 HTML 文件 ====================
    tmp_html = tempfile.NamedTemporaryFile(suffix='.html', delete=False, mode='w', encoding='utf-8')
    tmp_html.write(html_content)
    tmp_html.close()
    
    # ==================== Playwright 导出 PDF ====================
    print(f"\n🖨️ 正在通过 Playwright 无头浏览器渲染并导出 PDF ...")
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        
        page.goto(f'file://{tmp_html.name}', wait_until='networkidle')
        page.wait_for_timeout(1000)  # 确保中文字体渲染完成
        
        # 主报表有15列，用横向 landscape 更合适
        # 但项目工作记录只有5列，纵向更好
        # 折中方案：统一 landscape，所有表格都能装下
        page.pdf(
            path=pdf_path,
            format='A4',
            landscape=True,  # 横向以容纳主报表的15列
            print_background=True,  # 保留背景色
            margin={
                'top': '15mm',
                'bottom': '15mm',
                'left': '10mm',
                'right': '10mm'
            }
        )
        
        browser.close()
    
    # 清理临时文件
    os.unlink(tmp_html.name)
    
    file_size = os.path.getsize(pdf_path)
    size_str = f"{file_size/1024:.0f}KB" if file_size < 1024*1024 else f"{file_size/1024/1024:.1f}MB"
    
    print(f"\n🎉🎉 PDF 导出成功！")
    print(f"   📁 路径: {pdf_path}")
    print(f"   📏 大小: {size_str}")
    print(f"   📄 包含: {total_sheets} 个 Sheet 页")


def find_latest_excel():
    """自动找到 downloads 目录下最新的 Excel 文件"""
    download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'downloads')
    xlsx_files = glob.glob(os.path.join(download_dir, '**', '*.xlsx'), recursive=True)
    if not xlsx_files:
        print("❌ downloads 目录下没有找到 .xlsx 文件")
        return None
    # 按修改时间排序，取最新
    latest = max(xlsx_files, key=os.path.getmtime)
    return latest


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = find_latest_excel()
    
    if excel_file:
        excel_to_pdf(excel_file)
